"""Convert pathology-specific correlation CSVs into a marker-per-sheet workbook.

Each input CSV must be a square correlation matrix: the first row lists marker
names and the first column lists the same markers.  One output sheet is made
for each marker.  Its rows are pathologies and its columns are correlated
markers.
"""

from __future__ import annotations

import argparse
import csv
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_PATTERN = "*correlaitons_by-pathology_corr.csv"
HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def pathology_name(csv_path: Path) -> str:
    """Derive a readable pathology label from the standard correlation filename."""
    suffix = "_epi-log2-correlaitons_by-pathology_corr.csv"
    name = csv_path.name
    if name.endswith(suffix):
        return name[: -len(suffix)].replace("_", " ")
    return csv_path.stem.replace("_", " ")


def read_matrix(csv_path: Path) -> tuple[list[str], dict[str, dict[str, float]]]:
    """Read and validate one square, labeled correlation matrix."""
    with csv_path.open(newline="", encoding="utf-8-sig") as handle:
        rows = list(csv.reader(handle))

    if len(rows) < 2 or len(rows[0]) < 2:
        raise ValueError(f"{csv_path.name} is not a labeled correlation matrix.")

    markers = rows[0][1:]
    if not markers or any(not marker for marker in markers):
        raise ValueError(f"{csv_path.name} has empty marker headers.")

    matrix: dict[str, dict[str, float]] = {}
    for row in rows[1:]:
        if len(row) != len(markers) + 1:
            raise ValueError(f"{csv_path.name} has inconsistent row lengths.")
        source_marker = row[0]
        if source_marker not in markers:
            raise ValueError(f"{csv_path.name} has unexpected row marker {source_marker!r}.")
        matrix[source_marker] = dict(zip(markers, (float(value) for value in row[1:])))

    if set(matrix) != set(markers):
        raise ValueError(f"{csv_path.name} does not contain one row for every marker.")
    return markers, matrix


def safe_sheet_title(marker: str, used_titles: set[str]) -> str:
    """Make an Excel-compatible, unique sheet title while retaining the marker name."""
    base = re.sub(r"[\\/*?:\[\]]", "_", marker.rstrip("_")).strip() or "marker"
    base = base[:31]
    title = base
    index = 2
    while title in used_titles:
        suffix = f"_{index}"
        title = f"{base[: 31 - len(suffix)]}{suffix}"
        index += 1
    used_titles.add(title)
    return title


def build_workbook(input_dir: Path, output_path: Path, pattern: str) -> tuple[int, int]:
    csv_paths = sorted(input_dir.glob(pattern), key=lambda path: pathology_name(path).casefold())
    if not csv_paths:
        raise FileNotFoundError(f"No files matching {pattern!r} in {input_dir}")

    markers, first_matrix = read_matrix(csv_paths[0])
    matrices = [(pathology_name(csv_paths[0]), first_matrix)]
    expected_markers = set(markers)
    for csv_path in csv_paths[1:]:
        current_markers, matrix = read_matrix(csv_path)
        if set(current_markers) != expected_markers:
            raise ValueError(f"{csv_path.name} has a different set of markers.")
        matrices.append((pathology_name(csv_path), matrix))

    workbook = Workbook()
    workbook.remove(workbook.active)
    used_titles: set[str] = set()
    # Ecad is the requested first sheet; keep the source order for all others.
    sheet_markers = sorted(markers, key=lambda marker: (marker != "Ecad_", markers.index(marker)))
    for marker in sheet_markers:
        worksheet = workbook.create_sheet(safe_sheet_title(marker, used_titles))
        worksheet.append([marker.rstrip("_"), *markers])
        for pathology, matrix in matrices:
            worksheet.append([pathology, *(matrix[marker][partner] for partner in markers)])

        worksheet.freeze_panes = "B2"
        worksheet.auto_filter.ref = worksheet.dimensions
        worksheet.column_dimensions["A"].width = max(14, max(len(pathology) for pathology, _ in matrices) + 2)
        for column in range(2, len(markers) + 2):
            worksheet.column_dimensions[get_column_letter(column)].width = 13
        for cell in worksheet[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        for row in worksheet.iter_rows(min_row=2, min_col=2):
            for cell in row:
                cell.number_format = "0.000"

    workbook.save(output_path)
    return len(matrices), len(markers)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input_dir", type=Path, help="Folder containing correlation-matrix CSVs.")
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Workbook to create (default: marker_specific_correlations.xlsx in input_dir).",
    )
    parser.add_argument(
        "--pattern",
        default=DEFAULT_PATTERN,
        help=f"Input filename glob (default: {DEFAULT_PATTERN}).",
    )
    args = parser.parse_args()

    input_dir = args.input_dir.resolve()
    output_path = (args.output or input_dir / "marker_specific_correlations.xlsx").resolve()
    if output_path.parent != input_dir and not output_path.parent.exists():
        raise FileNotFoundError(f"Output directory does not exist: {output_path.parent}")
    pathology_count, marker_count = build_workbook(input_dir, output_path, args.pattern)
    print(f"Created {output_path} with {marker_count} marker sheets from {pathology_count} pathologies.")


if __name__ == "__main__":
    main()
